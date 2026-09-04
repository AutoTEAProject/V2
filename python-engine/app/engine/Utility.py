import math
from copy import deepcopy
from enums import EquipLen
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from data import HeaterParam, HeatExchangerParam, CompressorParam, reactorParam, ReactParam, atomicWeight, utilityCostData, equipmentConfig, HTX_CAPACITY_PARAM, CEPCI_BASE, CEPCI_CURRENT
from openpyxl import load_workbook


def checkType(name): 
	length = len(name)
	
	# 임시 예외처리
	if (name == "OLINHEA"):
		return "HTX"
	for i in range(0, length):
		# Mixer part
		if (length - i >= EquipLen.MIXER and name[i:i + EquipLen.MIXER] == "MIXER"):
			return "MIX"
		elif (length - i >= EquipLen.SSPLIT and name[i:i + EquipLen.SSPLIT] == "SSPLIT"):
			return "MIX"
		elif (length - i >= EquipLen.FSPLIT and name[i:i + EquipLen.FSPLIT] == "FSPLIT"):
			return "MIX"
		# Seperate part
		elif (length - i >= EquipLen.FLASH2 and name[i:i + EquipLen.FLASH2] == "FLASH2"):
			return "FLASH"
		elif (length - i >= EquipLen.FLASH3 and name[i:i + EquipLen.FLASH3] == "FLASH3"):
			return "FLASH"
		elif (length - i >= EquipLen.SEP and name[i:i + EquipLen.SEP] == "SEP"):
			return "SEP"
		elif (length - i >= EquipLen.SEP2 and name[i:i + EquipLen.SEP2] == "SEP2"):
			return "SEP"
		# Exchanger part
		elif (length - i >= EquipLen.HEATER and name[i:i + EquipLen.HEATER] == "HEATER"):
			return "HTX"
		elif (length - i >= EquipLen.HEATX and name[i:i + EquipLen.HEATX] == "HEATX"):
			return "HEX"
		elif (length - i >= EquipLen.MHEATX and name[i:i + EquipLen.MHEATX] == "MHEATX"):
			return "HEX"
		elif (length - i >= EquipLen.HXFLUX and name[i:i + EquipLen.HXFLUX] == "HXFLUX"):
			return "HEX"
		# Column part
		elif (length - i >= EquipLen.DSTWU and name[i:i + EquipLen.DSTWU] == "DSTWU"):
			return "DIST"
		elif (length - i >= EquipLen.DISTL and name[i:i + EquipLen.DISTL] == "DISTL"):
			return "DIST"
		elif (length - i >= EquipLen.RADFRAC and name[i:i + EquipLen.RADFRAC] == "RADFRAC"):
			return "DIST"
		elif (length - i >= EquipLen.PETROFRAC and name[i:i + EquipLen.PETROFRAC] == "PETROFRAC"):
			return "DIST"
		# Reactor Part
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RSTOIC"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RYIELD"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "REQUIL"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RGIBBS"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RCSTR"):
			return "REACT"
		elif (length - i >= EquipLen.RSTOIC and name[i:i + EquipLen.RSTOIC] == "RPLUG"):
			return "REACT"
		# Pressure Changer Part
		elif (length - i >= EquipLen.COMPR and name[i:i + EquipLen.COMPR] == "COMPR"):
			return "COMP"
		elif (length - i >= EquipLen.COMP and name[i:i + EquipLen.COMP] == "COMP"):
			return "COMP"
		elif (length - i >= EquipLen.MCOMPR and name[i:i + EquipLen.MCOMPR] == "MCOMPR"):
			return "COMP"
		# 아래 두 개는 계산 아스펜에서 해줘서 원래 벨브 처리 어떻게 했는지 확인필요
		elif (length - i >= EquipLen.PUMP and name[i:i + EquipLen.PUMP] == "PUMP"):
			return "VALVE"
		elif (length - i >= EquipLen.VALVE and name[i:i + EquipLen.VALVE] == "VALVE"):
			return "VALVE"
		# Solid Seperator Part -> aspen에서 계산해줌
		elif (length - i >= EquipLen.CYCLONE and name[i:i + EquipLen.CYCLONE] == "CYCLONE"):
			return "DIST"
		elif (length - i >= EquipLen.HYCYC and name[i:i + EquipLen.HYCYC] == "HYCYC"):
			return "DIST"
	# print("Type을 확인할 수 없는 장비가 있습니다. : " + name)
	return "ETC"

def selectedFormulaNames(key, paramDict):
	"""
	장치 key에 대해 체크박스로 선택된 후보 수식 이름들을 돌려준다.
	설정이 없는 장치(예: 파싱 스냅샷에 없던 예외 상황)는 방어적으로 해당 타입의 모든 후보를 계산한다.
	"""
	names = equipmentConfig.get("equipment", {}).get(key, {}).get("selectedFormulas")
	if names:
		return [name for name in names if name in paramDict]
	return list(paramDict.keys())

def powerLawCost(params, capacity, capacityUnit):
	"""
	HEX/HTX 공통 상관식: EQUIPMENT COST = 10^(K1+K2+K3) * (Capacity/10)^0.6 * (CEPCI_CURRENT/CEPCI_BASE)
	계산 결과뿐 아니라 계산에 쓰인 입력값도 같이 돌려준다. "FormulaKind"는 결과 엑셀(printout)이
	EQUIPMENT COST 셀에 어떤 모양의 수식을 넣어야 하는지 판단하는 데만 쓰는 내부 표시다.
	"""
	cost = ((10 ** (params["K1"] + params["K2"] + params["K3"])) * (capacity / 10) ** 0.6) * (CEPCI_CURRENT / CEPCI_BASE)
	entry = deepcopy(params)
	entry["Capacity"] = capacity
	entry["Capacity Unit"] = capacityUnit
	entry["CEPCI(base)"] = CEPCI_BASE
	entry["CEPCI(current)"] = CEPCI_CURRENT
	entry["EQUIPMENT COST"] = cost
	entry["FormulaKind"] = "powerLaw"
	return entry

def logPolynomialCost(params, capacity, capacityUnit):
	"""COMP 상관식: EQUIPMENT COST = 10^(K1 + K2*log10(Capacity) + K3*log10(Capacity)^2) * (CEPCI_CURRENT/CEPCI_BASE)"""
	logCapacity = math.log(capacity, 10)
	cost = (10 ** (params["K1"] + params["K2"] * logCapacity + params["K3"] * (logCapacity ** 2))) * (CEPCI_CURRENT / CEPCI_BASE)
	entry = deepcopy(params)
	entry["Capacity"] = capacity
	entry["Capacity Unit"] = capacityUnit
	entry["CEPCI(base)"] = CEPCI_BASE
	entry["CEPCI(current)"] = CEPCI_CURRENT
	entry["EQUIPMENT COST"] = cost
	entry["FormulaKind"] = "logPoly"
	return entry

def normalizeName(name):
	# 이름이 .으로 구분되어 있으면 . 이후의 문자열을 반환, 아니면 원래 이름 반환
	if "." in name:
		return name.split(".")[-1]
	return name

def calMaterialWeight(material):
	atom = ""
	count = 0
	weight = 0
	for i in range(len(material)):
		if material[i].isalpha() and material[i].isupper():
			if atom == "":
				atom = material[i]
			else:
				if (count == 0):
					count += 1
				if atom in atomicWeight:
					weight += atomicWeight[atom] * count
				else :
					raise ValueError(f"{atom} is not in atomicWeight Data")
				count = 0
				atom = material[i]
		elif material[i].isalpha() and material[i].islower():
			atom += material[i]
		elif material[i].isdigit():
			count = count * 10 + int(material[i])
	if (atom != ""):
		if (count == 0):
			count += 1
		weight += atomicWeight[atom] * count
	return (weight)
	

# cost = {} # 2차원 딕셔너리로 "이름" : {딕셔너리} 이렇게 저장하고 각 유닛 종류별 인자와 계산 결과를 출력한다.
def calEquipmentCost(inputData, cost, utility, exceptCapacity): #react도 추가해야함.
	for key in inputData:
		temp = {}
		type = inputData[key]["Type"]
		exceptflag = False
 		# 여기서 이제 cost의 값들을 하나씩 이름, 인자 순으로 저장해야함.
		'''
		1. EQUIPMENT COST
		- HEX, HTX : ((10^(K1+K2+K3))*(Capacity / 10)^(0.6)) * (CEPCI(June 2024) / CEPCI(Sept 2001))
		- COMP : (10^(K1 + K2 * log(Capacity) + K3 * ((log(Capacity))^2))) * (CEPCI(June 2024) / CEPCI(Sept 2001))
		2. C_BM 
		- HEX, HTX : EQUIPMENT COST * (B1 + B2*FM)
		'''

		# utility 사용량이 아예 파싱되지 않은 장치(.rep에 해당 유틸리티 섹션이 없는 경우)도 있을 수 있어서
		# utility[key]를 바로 인덱싱하지 않고 매 반복마다 새로 조회한다(예전 값이 다음 장치로 새는 버그도 같이 방지).
		utilityKey = None
		equipmentUtility = utility.get(key, {})

		if (type == "HTX"):
			if ("HOT UTILITY[kW]" in equipmentUtility):
				utilityKey = "HOT UTILITY[kW]"
			elif ("ELECTRICITY UTILITY[kW]" in equipmentUtility):
				utilityKey = "ELECTRICITY UTILITY[kW]"
			elif ("COOLING UTILITY[kg/hr]" in equipmentUtility):
				utilityKey = "COOLING UTILITY[kg/hr]"
			if utilityKey:
				capacity = equipmentUtility[utilityKey]
		elif (type == "HEX"):
			capacity = inputData[key]["HeatTransferArea"]
		elif (type == "COMP"):
			capacity = inputData[key]["DriverPower"]
		if (type == "HTX" or checkType(key) == "HTX"):
			if utilityKey is None:
				print(key, ": HTX 장치인데 utility 사용량(HOT/ELECTRICITY/COOLING)이 파싱되지 않아 장치비를 계산할 수 없습니다.")
			else:
				capacity = equipmentUtility[utilityKey] / HTX_CAPACITY_PARAM
				for formulaName in selectedFormulaNames(key, HeaterParam):
					temp[formulaName] = powerLawCost(HeaterParam[formulaName], capacity, "kW")
		elif (type == "HEX"):
			for formulaName in selectedFormulaNames(key, HeatExchangerParam):
				temp[formulaName] = powerLawCost(HeatExchangerParam[formulaName], capacity, "sqm")
		elif (checkType(key) == "COMP"):
			if (inputData[key]["DriverPower"] == 0):
				print("Driver Power이 0인 Compressor가 있습니다. : " + key)
				continue
				#error
				# 나중에 여기에 에러 처리 코드 넣기
			capacity = inputData[key]["DriverPower"]
			for formulaName in selectedFormulaNames(key, CompressorParam):
				temp[formulaName] = logPolynomialCost(CompressorParam[formulaName], capacity, "kW")
		elif (type == "REACT"):
			if key in reactorParam:
				p = reactorParam[key]
				reactCost = (p["Equipment Cost"] * (p["Capacity_parsed KW"] / p["Capacity [KW]"]) ** p["Scaling Factor"]) * p["Additional Param"] * (p["Cepci_recent"] / p["Cepci_before"])
				temp["input"] = deepcopy(p)
				temp["input"]["EQUIPMENT COST"] = reactCost
				temp["input"]["FormulaKind"] = "reactScaling"
			else:
				temp["input"] = deepcopy(ReactParam["Nan"])
				temp["input"]["EQUIPMENT COST"] = 0
		# 여기는 이미 가격 계산 되어있으면 계산 안 하는 부분
		if (key not in exceptCapacity and inputData[key]["EquipmentCost"] != 0 and checkType(key) not in ["HTX", "HEX", "COMP"]):
			# print("엑셀에 이미 가격이 입력되어 있습니다. 계산된 가격과 비교해서 확인해주세요. : " + key)
			# print(checkType(key))
			temp["ATEA"] = {}
			temp["ATEA"]["EQUIPMENT COST"] = inputData[key]["EquipmentCost"]
		# if (key == "CO2H2MC.1MCCOMP2"):
		# 	print("CO2H2MC.1MCCOMP2 찾음")
		# 	print(temp)
		# 	print(type)
		# 	print(checkType(key))
		cost[key] = deepcopy(temp)
		# print(cost[key])
	 
# 이제 여기서 Capacity 값은 각 모듈별로 파싱해서 저장해둬야함.
def safe_df(df):
    # 인덱스가 RangeIndex(0,1,2,...)가 아니거나, 인덱스 이름이 있다면 컬럼으로 복구
    if not isinstance(df.index, pd.RangeIndex) or df.index.name is not None:
        return df.reset_index()
    return df

def write_equipment_cost_block(ws, equipmentName, formulas):
    """
    장치 하나(equipmentName)에 대해 후보 수식(formulas: 수식 이름 -> 필드 dict)을 표로 적는다.
    "EQUIPMENT COST" 칸은 그냥 계산된 숫자를 적는 게 아니라, 같은 행의 K1/K2/K3/Capacity/CEPCI
    같은 셀들을 참조하는 실제 엑셀 수식(=...)으로 적어서, 엑셀에서 그 셀을 클릭하면 어떻게
    계산됐는지 바로 보이고 값을 바꾸면 다시 계산되게 한다.
    """
    ws.append([equipmentName])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="004085")

    fieldOrder = []
    for fields in formulas.values():
        for field in fields:
            if field not in fieldOrder:
                fieldOrder.append(field)
    ws.append(["formula"] + fieldOrder)

    for formulaName, fields in formulas.items():
        kind = fields.get("FormulaKind")
        rowValues = [formulaName] + [fields.get(field) for field in fieldOrder]
        ws.append(rowValues)
        row = ws.max_row
        colIndex = {field: i for i, field in enumerate(fieldOrder, start=2)}

        def coord(field):
            return ws.cell(row=row, column=colIndex[field]).coordinate

        if kind == "powerLaw":
            formula = (
                f"=10^({coord('K1')}+{coord('K2')}+{coord('K3')})"
                f"*({coord('Capacity')}/10)^0.6"
                f"*({coord('CEPCI(current)')}/{coord('CEPCI(base)')})"
            )
            ws.cell(row=row, column=colIndex["EQUIPMENT COST"]).value = formula
        elif kind == "logPoly":
            formula = (
                f"=10^({coord('K1')}+{coord('K2')}*LOG10({coord('Capacity')})"
                f"+{coord('K3')}*LOG10({coord('Capacity')})^2)"
                f"*({coord('CEPCI(current)')}/{coord('CEPCI(base)')})"
            )
            ws.cell(row=row, column=colIndex["EQUIPMENT COST"]).value = formula
        elif kind == "reactScaling":
            formula = (
                f"={coord('Equipment Cost')}"
                f"*({coord('Capacity_parsed KW')}/{coord('Capacity [KW]')})^{coord('Scaling Factor')}"
                f"*{coord('Additional Param')}*({coord('Cepci_recent')}/{coord('Cepci_before')})"
            )
            ws.cell(row=row, column=colIndex["EQUIPMENT COST"]).value = formula
        # kind가 없으면(ATEA, 혹은 REACT 파라미터가 없어 0으로 대체된 경우) Aspen이 준 값이거나
        # 계산할 수식이 아예 없는 경우라서 EQUIPMENT COST는 그냥 숫자 그대로 둔다.
    ws.append([])

def autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

def printout(inputData, cost, utility, CAPEX, OPEX, profitAnalysis):
    
    columns = ["Name","EquipmentCost","InstalledCost","EquipmentWeight",
           "InstalledWeight","UtilityCost","HeatTransferArea","DriverPower"]
    wb = Workbook()
    ws = wb.active
    ws.title = "parse"
    # # 데이터 추가
    ws.append(columns)

# 2️⃣ 데이터 추가
    for name, row in inputData.items():
        ws.append([name] + [row[col] for col in columns[1:]])  # 첫 열은 name, 나머지는 columns 순서대로
    autofit(ws)

    # 시트 2: UTILITY
    ws2 = wb.create_sheet("UTILITY")
    for r in dataframe_to_rows(pd.DataFrame(utility), header=True): ws2.append(r)
    autofit(ws2)
    
    # 시트 3: CAPCOST (블럭 스타일). EQUIPMENT COST 칸은 실제 엑셀 수식으로 적는다(write_equipment_cost_block 참고).
    ws3 = wb.create_sheet("Equipment Cost")
    for k, v in cost.items():
        write_equipment_cost_block(ws3, k, v)
    autofit(ws3)

    ws4 = wb.create_sheet("CAPEX")

    for key, values in CAPEX.items():
        ws4.append([key] + values)

    autofit(ws4)
    
    # 시트 5: OPEX
    ws5 = wb.create_sheet("OPEX")
    for key, values in OPEX.items():
        ws5.append([key] + values)
    autofit(ws5)
    
    ws6 = wb.create_sheet("Profitability Analysis")
    opex_df = pd.DataFrame([profitAnalysis]).T  if isinstance(profitAnalysis, dict) else pd.DataFrame(profitAnalysis).T 
    for r in dataframe_to_rows(opex_df, header=False):
        ws6.append(r)
    autofit(ws6)
    
    # 마지막에 한 번만 저장
    wb.save("output.xlsx")

def parseReact(excelReactorData):
	filename = "./input/MaterialData.xlsx"
	df = pd.read_excel(io = filename, sheet_name='Reactor', header=1, engine='openpyxl')
	length = len(df)
	for i in range(length):
		name = df.iat[i, 1]
		cost = df.iat[i, 2]
		excelReactorData[name] = cost

def inputReactorName(inputData, cost):
	filename = "./input/MaterialData.xlsx"

	# 기존 엑셀 로드
	wb = load_workbook(filename)
	ws = wb["Reactor"]

	# pandas로 DF 로드 (기존 데이터 확인용)
	df = pd.read_excel(filename, sheet_name='Reactor', header=1, engine='openpyxl')

	# name 컬럼이 실제 몇 번째인지 찾기
	name_col_index = list(df.columns).index("Reactor Name") + 1  # excel column index (+1 필요)

	current_row = 3  # 데이터 시작 row

	for key in cost:
		if inputData[key]["Type"] == "REACT":
			ws.cell(row=current_row, column=name_col_index).value = key
			current_row += 1

	wb.save(filename)
			

def inputRTX(inputData, cost):
	excelReactorData = {};
	parseReact(excelReactorData)
	for key in cost:
		if (inputData[key]["Type"] == "REACT"):
			if key in excelReactorData:
				cost[key]["input"]["EQUIPMENT COST"] = excelReactorData[key]
			else:
				inputReactorName(inputData, cost)
				raise Exception("reactor의 가격을 엑셀 시트에 이름에 맞게 입력해주세요")
